#!/usr/bin/env python3
"""Parse Lego.xlsx into products.json with normalized columns.

Usage: run from the /Users/patrick/lego directory where Lego.xlsx was copied.
"""
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as e:
    print("openpyxl is required. Please run: python3 -m pip install --user openpyxl")
    raise

HERE = Path(__file__).resolve().parent
XLSX = HERE / 'Lego.xlsx'
OUT = HERE / 'products.json'

def normalize_row(row_dict):
    # lower keys
    map_lower = { (k.strip().lower() if isinstance(k,str) else k): v for k,v in row_dict.items() }
    def g(*keys):
        for k in keys:
            if k in map_lower and map_lower[k] is not None:
                return map_lower[k]
        return ''

    idv = g('id','sku','part','item')
    title = str(g('title','name'))
    description = str(g('description','desc'))
    # prefer Cost, then price, then Total
    price_raw = g('cost','price','unit_price','total')
    # print(price_raw)
    try:
        price = float(price_raw)
    except Exception:
        # try to strip currency
        try:
            price = float(str(price_raw).replace('$','').replace(',',''))
        except Exception:
            price = 0.0
    qty_raw = g('quantity','qty','count')
    try:
        quantity = int(qty_raw)
    except Exception:
        try:
            quantity = int(float(str(qty_raw)))
        except Exception:
            quantity = 0
    # add 1 to each set's quantity per user request
    # quantity parsed from the sheet (do not add here); we'll compute totals by counting occurrences
    # original MSRP / list price if present
    msrp_raw = g('msrp','rrp','list_price','retail','recommended_retail_price')
    try:
        msrp = float(msrp_raw)
    except Exception:
        try:
            msrp = float(str(msrp_raw).replace('$','').replace(',',''))
        except Exception:
            msrp = 0.0

    imageUrl = str(g('imageurl','image','img'))
    condition = str(g('condition'))

    if not idv:
        idv = title[:50] or 'unknown'

    # normalize numeric-looking ids like '10267.0' -> '10267'
    id_str = str(idv)
    if id_str.endswith('.0'):
        id_str = id_str[:-2]

    return {
        'id': id_str,
        'title': title,
        'description': description,
        'price': price,
        'msrp': msrp,
        'quantity': quantity,
        'imageUrl': imageUrl,
        'condition': condition
    }

def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found. Make sure the file was copied into the project.")
        sys.exit(2)

    wb = load_workbook(filename=str(XLSX), data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("No rows found in the spreadsheet.")
        sys.exit(1)

    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    data_rows = rows[1:]

    parsed = []
    for r in data_rows:
        row_dict = { header[i] if i < len(header) else f'col{i}': r[i] if i < len(r) else '' for i in range(len(header)) }
        parsed.append(normalize_row(row_dict))

    # Aggregate products by `id` and count how many times each set appears in the sheet
    products_by_id = {}
    for p in parsed:
        pid = p['id']
        if pid in products_by_id:
            products_by_id[pid]['quantity'] = products_by_id[pid].get('quantity', 0) + 1
            # prefer a non-empty title/price/image if current entry lacks it
            if not products_by_id[pid].get('title') and p.get('title'):
                products_by_id[pid]['title'] = p.get('title')
            if (not products_by_id[pid].get('price') or products_by_id[pid].get('price') == 0.0) and p.get('price'):
                products_by_id[pid]['price'] = p.get('price')
            if not products_by_id[pid].get('imageUrl') and p.get('imageUrl'):
                products_by_id[pid]['imageUrl'] = p.get('imageUrl')
        else:
            copy = p.copy()
            # initialize quantity to 1 for the first occurrence
            copy['quantity'] = 1
            products_by_id[pid] = copy

    final_products = list(products_by_id.values())

    # write JSON
    with OUT.open('w', encoding='utf-8') as f:
        json.dump(final_products, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(final_products)} unique products to {OUT}")
    print("Preview (first 10):")
    for p in final_products[:10]:
        msrp_str = f" msrp=${p['msrp']}" if p.get('msrp') else ''
        print(f"- {p['id']} | {p['title']} | ${p['price']} | qty={p['quantity']}{msrp_str}")

if __name__ == '__main__':
    main()
